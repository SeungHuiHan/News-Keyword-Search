from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# 네이버 검색창에 검색하고 싶은 뉴스의 키워드로 검색한다.
# 알고 싶은 키워드를 크롤링하여 키워드와 뉴스 10개를 크롤링하여 엑셀에 정리 해주는 프로그램


Keyword = input("검색할 뉴스의 키워드를 말해주세요: ")
prefer = int(input("뉴스를 관련도순으로 볼려면 1 최신순으로 볼려면 2를 입력하세요: "))

# WebDriver 설정
browser = webdriver.Chrome("./chromedriver.exe") #chromedriver가 있는 파일 경로 입력
browser.get('https://search.naver.com/search.naver?ssc=tab.news.all&where=news&sm=tab_jum&query='+Keyword)
browser.maximize_window()

# 기사 정렬 방식 선택
prefer_lineup = browser.find_element(By.XPATH, '//*[@id="snb"]').find_element(By.CSS_SELECTOR, '.bx.lineup')

if prefer == 1:
    browser.find_element(By.CLASS_NAME, 'mod_list_option_filter').find_element(By.LINK_TEXT, "관련도순").click() 
elif prefer == 2:
    browser.find_element(By.CLASS_NAME, 'mod_list_option_filter').find_element(By.LINK_TEXT, "최신순").click() 

browser.implicitly_wait(2)

# 지면기사 선택
browser.find_element(By.XPATH, '//*[@id="snb"]/div[1]/div/div[2]/a').send_keys(Keys.ENTER)
browser.find_element(By.CLASS_NAME, 'bx.view').find_element(By.LINK_TEXT, "지면기사").send_keys(Keys.ENTER)

# BeautifulSoup으로 파싱
soup = BeautifulSoup(browser.page_source, 'html.parser')
news_list = soup.find(class_="list_news _infinite_list")
articles = news_list.select('.list_news > li') # 뉴스 article select

# 데이터프레임 생성
data = []
for article in articles[:10]:
    a_elements = article.select('div.news_area > div.news_contents > a') # 제목
    if len(a_elements) == 1:
        title = a_elements[0].text
    else:
        title = a_elements[1].text

    press_element = article.select_one('div.news_area > div.news_info > div.info_group > a')
    if press_element:
        press_text = press_element.text  # 언론사 이름과 '언론사 선정' 텍스트 가져오기
        if '언론사 선정' in press_text:
            press = press_text.replace('언론사 선정', '')  # '언론사 선정' 텍스트 삭제
        else:
            press = press_text

    upload_time = article.select_one('div.news_area > div.news_info > div.info_group > span:nth-of-type(2)').text # 게시 시간
    url = article.select_one('div.news_area > div.news_contents > a')['href']  # 링크URL 

    data.append([title, press, upload_time, url])

# 데이터프레임 생성
df = pd.DataFrame(data, columns=['제목', '언론사', '게시 시간', '링크'])

# 엑셀 파일 저장
now = datetime.now().strftime("%Y년 %m월 %d일 %H시 %M분 %S초")
excel_file_name = f"{Keyword} {'관련도순' if prefer == 1 else '최신순'} 뉴스 ({now}).xlsx"
df.to_excel(excel_file_name, index=False)

# 엑셀 파일에 하이퍼링크 추가
wb = load_workbook(excel_file_name)
ws = wb.active

# 제목 열의 너비 조정
title_column_width = max(len(title) for title in df['제목'])
ws.column_dimensions['A'].width = title_column_width * 1.5  # 너비 조정

# 언론사 열의 너비 조정
press_column_width = max(len(press) for press in df['언론사'])
ws.column_dimensions['B'].width = press_column_width * 2.0  # 너비 조정

# 게시 시간 열의 너비 조정
upload_time_column_width = max(len(upload_time) for upload_time in df['게시 시간'])
ws.column_dimensions['C'].width = upload_time_column_width * 1.5  # 너비 조정

for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=4, max_col=4):
    for cell in row:
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"
        cell.value = "링크"

wb.save(excel_file_name)
wb.close()

print(f"데이터를 {excel_file_name}에 저장했습니다.")

# 브라우저 종료
browser.quit()

# 저장한 엑셀 파일 열기
#os.system(f"start {excel_file_name}")
